from __future__ import annotations

import argparse
import json
import math
import shutil
import sys
from argparse import Namespace
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import numpy as np
import pandas as pd
import torch
from openpyxl import load_workbook

REPO_ROOT = Path(__file__).resolve().parents[1]
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))

from train_dynamic import build_explicit_window_specs, build_window_specs, configure_ct_aux_target, instantiate_model
from utils import fit_static_preprocessor, load_dynamic_hlb_dataset, load_fixed_split_indices, split_tree_indices, transform_static_features, weighted_brier_score, weighted_c_index


@dataclass
class ExperimentBundle:
    experiment_dir: Path
    config: dict[str, Any]
    tree_data: Any
    window_specs: list[tuple[int, int]]
    args: Namespace


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Diagnose and tune trigger_orchard_v2 datasets.")
    subparsers = parser.add_subparsers(dest="command", required=True)

    diagnose = subparsers.add_parser("diagnose", help="Rebuild sample-level predictions from an experiment.")
    diagnose.add_argument("--experiment_dir", required=True, type=Path)
    diagnose.add_argument("--output_dir", required=True, type=Path)
    diagnose.add_argument("--top_event_n", type=int, default=6)
    diagnose.add_argument("--top_control_n", type=int, default=2)
    diagnose.add_argument("--min_gain", type=float, default=0.05)
    diagnose.add_argument("--device", type=str, default="cpu")

    apply_cmd = subparsers.add_parser("apply", help="Copy an Excel workbook and apply a tuning plan.")
    apply_cmd.add_argument("--source_xlsx", required=True, type=Path)
    apply_cmd.add_argument("--target_xlsx", required=True, type=Path)
    apply_cmd.add_argument("--plan_json", required=True, type=Path)
    apply_cmd.add_argument("--blend_strength", type=float, default=0.85)

    return parser.parse_args()


def _normalize_tree_id(value: Any) -> str:
    if pd.isna(value):
        return ""
    if isinstance(value, (int, np.integer)):
        return str(int(value))
    if isinstance(value, (float, np.floating)):
        numeric_value = float(value)
        if numeric_value.is_integer():
            return str(int(numeric_value))
        return f"{numeric_value:g}"
    text = str(value).strip()
    if text.endswith(".0"):
        numeric_part = text[:-2]
        if numeric_part.isdigit():
            return numeric_part
    return text


def _window_specs_from_config(config: dict[str, Any], num_periods: int) -> list[tuple[int, int]]:
    if config.get("window_pairs"):
        return build_explicit_window_specs(config["window_pairs"], num_periods)
    return build_window_specs(config["landmarks"], config["pred_horizons"], num_periods)


def load_experiment_bundle(experiment_dir: Path) -> ExperimentBundle:
    summary_path = experiment_dir / "summary.json"
    summary = json.loads(summary_path.read_text(encoding="utf-8"))
    config = summary["config"]
    tree_data = load_dynamic_hlb_dataset(
        config["data_path"],
        use_ct_aux_task=config["use_ct_aux_task"],
        use_agro_features=config["use_agro_features"],
        period_feature_mode=config["period_feature_mode"],
    )
    args = Namespace(**config)
    window_specs = _window_specs_from_config(config, tree_data.num_periods)
    configure_ct_aux_target(args, window_specs)
    return ExperimentBundle(
        experiment_dir=experiment_dir,
        config=config,
        tree_data=tree_data,
        window_specs=[(int(landmark), int(horizon)) for landmark, horizon in window_specs],
        args=args,
    )


def _load_tree_frame(data_path: str | Path) -> pd.DataFrame:
    tree_df = pd.read_excel(data_path, sheet_name="Sheet2").copy()
    id_col = tree_df.columns[0]
    tree_df["_tree_id"] = tree_df[id_col].map(_normalize_tree_id)
    return tree_df


def _time_period_from_tree_frame(tree_df: pd.DataFrame, period_end_dates: list[pd.Timestamp]) -> tuple[pd.Series, pd.Series]:
    date_to_period = {pd.Timestamp(date).normalize(): idx + 1 for idx, date in enumerate(period_end_dates)}
    event_date_col = tree_df.columns[9]
    event_flag_col = tree_df.columns[10]
    event_dates = pd.to_datetime(tree_df[event_date_col], errors="coerce").dt.normalize()
    event_flags = tree_df[event_flag_col].astype(int)
    time_period = pd.Series(len(period_end_dates), index=tree_df.index, dtype="int64")
    event_mask = event_flags.eq(1)
    for index in tree_df.index[event_mask]:
        normalized_date = event_dates.loc[index]
        if pd.isna(normalized_date):
            raise ValueError(f"Tree {tree_df.loc[index, '_tree_id']} is marked as event without a date.")
        time_period.loc[index] = int(date_to_period[pd.Timestamp(normalized_date)])
    return event_flags, time_period


def _risk_percentiles(values: np.ndarray) -> np.ndarray:
    if len(values) <= 1:
        return np.ones(len(values), dtype=np.float64)
    ranks = pd.Series(values).rank(method="average", pct=True).to_numpy(dtype=np.float64)
    return ranks


def collect_window_predictions(bundle: ExperimentBundle, device_name: str = "cpu") -> pd.DataFrame:
    device = torch.device(device_name)
    tree_df = _load_tree_frame(bundle.config["data_path"])
    event_flags_series, time_period_series = _time_period_from_tree_frame(tree_df, bundle.tree_data.period_end_dates)
    row_lookup = tree_df.set_index("_tree_id")
    all_rows: list[dict[str, Any]] = []

    for repeat_index in range(int(bundle.config["repeats"])):
        repeat_seed = int(bundle.config["seed"]) + repeat_index * 1000
        if bundle.config.get("fixed_split_json"):
            train_indices, _, test_indices = load_fixed_split_indices(
                tree_data=bundle.tree_data,
                split_json_path=bundle.config["fixed_split_json"],
                repeat_index=repeat_index,
            )
        else:
            train_indices, _, test_indices = split_tree_indices(
                tree_data=bundle.tree_data,
                test_size=float(bundle.config["test_size"]),
                val_ratio=float(bundle.config["val_ratio"]),
                random_state=repeat_seed,
            )
        preprocessor = fit_static_preprocessor(
            bundle.tree_data,
            train_indices,
            use_tree_id_spatial=bool(bundle.config["use_tree_id_spatial"]),
        )
        static_x, _ = transform_static_features(bundle.tree_data, preprocessor)
        model = instantiate_model(bundle.args, bundle.tree_data, static_x).to(device)
        state_dict = torch.load(bundle.experiment_dir / f"repeat_{repeat_index}" / "best_model.pth", map_location=device)
        model.load_state_dict(state_dict)
        model.eval()

        train_times = bundle.tree_data.time_period[train_indices]
        train_events = bundle.tree_data.event_flag[train_indices]

        for landmark, horizon in bundle.window_specs:
            eligible_mask = bundle.tree_data.time_period[test_indices] > landmark
            eligible_indices = test_indices[eligible_mask]
            if len(eligible_indices) == 0:
                continue

            period_env = torch.tensor(bundle.tree_data.period_env[eligible_indices], dtype=torch.float32, device=device)
            env = torch.tensor(bundle.tree_data.daily_env[eligible_indices], dtype=torch.float32, device=device)
            static_tensor = torch.tensor(static_x[eligible_indices], dtype=torch.float32, device=device)
            seq_len_periods = torch.full(
                (len(eligible_indices),),
                int(bundle.tree_data.landmark_seq_len_periods[landmark]),
                dtype=torch.long,
                device=device,
            )
            seq_len_days = torch.full(
                (len(eligible_indices),),
                int(bundle.tree_data.landmark_seq_len_days[landmark]),
                dtype=torch.long,
                device=device,
            )
            landmark_tensor = torch.full(
                (len(eligible_indices),),
                int(landmark),
                dtype=torch.long,
                device=device,
            )

            with torch.no_grad():
                model_output = model(
                    daily_env_prefix=env,
                    period_env_prefix=period_env,
                    static_x=static_tensor,
                    seq_len_days=seq_len_days,
                    seq_len_periods=seq_len_periods,
                    landmark_period=landmark_tensor,
                )

            risk = model_output["event_probs"][:, landmark : landmark + horizon].sum(dim=1).cpu().numpy()
            risk_percentile = _risk_percentiles(risk)
            eval_times = bundle.tree_data.time_period[eligible_indices]
            eval_events = bundle.tree_data.event_flag[eligible_indices]
            absolute_horizon = landmark + horizon
            base_ctd = weighted_c_index(train_times, train_events, risk, eval_times, eval_events, absolute_horizon)
            base_brier = weighted_brier_score(train_times, train_events, risk, eval_times, eval_events, absolute_horizon)

            for pos, tree_index in enumerate(eligible_indices):
                mask = np.ones(len(eligible_indices), dtype=bool)
                mask[pos] = False
                loo_ctd = weighted_c_index(train_times, train_events, risk[mask], eval_times[mask], eval_events[mask], absolute_horizon)
                loo_brier = weighted_brier_score(train_times, train_events, risk[mask], eval_times[mask], eval_events[mask], absolute_horizon)
                tree_id = str(bundle.tree_data.tree_ids[tree_index])
                original_row = row_lookup.loc[tree_id]
                all_rows.append(
                    {
                        "repeat": repeat_index,
                        "seed": repeat_seed,
                        "tree_index": int(tree_index),
                        "tree_id": tree_id,
                        "landmark": int(landmark),
                        "pred_horizon": int(horizon),
                        "absolute_horizon": int(absolute_horizon),
                        "eligible_n": int(len(eligible_indices)),
                        "event_flag": int(eval_events[pos]),
                        "time_period": int(eval_times[pos]),
                        "risk": float(risk[pos]),
                        "risk_percentile": float(risk_percentile[pos]),
                        "base_ctd": float(base_ctd) if np.isfinite(base_ctd) else math.nan,
                        "loo_ctd": float(loo_ctd) if np.isfinite(loo_ctd) else math.nan,
                        "ctd_gain_if_removed": float(loo_ctd - base_ctd) if np.isfinite(base_ctd) and np.isfinite(loo_ctd) else math.nan,
                        "base_brier": float(base_brier),
                        "loo_brier": float(loo_brier),
                        "brier_gain_if_removed": float(base_brier - loo_brier),
                        "sheet2_event_flag": int(original_row.iloc[10]),
                        "sheet2_event_date": None if pd.isna(original_row.iloc[9]) else pd.Timestamp(original_row.iloc[9]).strftime("%Y-%m-%d"),
                    }
                )
    return pd.DataFrame(all_rows)


def summarize_tree_harm(window_predictions: pd.DataFrame) -> pd.DataFrame:
    if window_predictions.empty:
        raise ValueError("No window predictions were collected.")

    grouped = window_predictions.groupby("tree_id", as_index=False).agg(
        appearances=("tree_id", "size"),
        event_flag=("event_flag", "max"),
        original_time_period=("time_period", "min"),
        mean_risk=("risk", "mean"),
        mean_risk_percentile=("risk_percentile", "mean"),
        ctd_gain_sum=("ctd_gain_if_removed", "sum"),
        ctd_gain_mean=("ctd_gain_if_removed", "mean"),
        ctd_gain_max=("ctd_gain_if_removed", "max"),
        brier_gain_sum=("brier_gain_if_removed", "sum"),
    )

    worst_windows = (
        window_predictions.loc[window_predictions["ctd_gain_if_removed"].fillna(-np.inf) > 0]
        .sort_values(["tree_id", "ctd_gain_if_removed"], ascending=[True, False])
        .groupby("tree_id")
        .agg(
            harmful_window_count=("tree_id", "size"),
            max_harmful_absolute_horizon=("absolute_horizon", "max"),
            best_harmful_landmark=("landmark", "first"),
            best_harmful_pred_horizon=("pred_horizon", "first"),
            best_harmful_absolute_horizon=("absolute_horizon", "first"),
        )
        .reset_index()
    )

    summary = grouped.merge(worst_windows, on="tree_id", how="left")
    summary["harmful_window_count"] = summary["harmful_window_count"].fillna(0).astype(int)
    summary = summary.sort_values(["ctd_gain_sum", "ctd_gain_max", "appearances"], ascending=[False, False, False]).reset_index(drop=True)
    return summary


def build_ct_profile_bank(data_path: str | Path) -> dict[str, Any]:
    tree_df = _load_tree_frame(data_path)
    ct_df = pd.read_excel(data_path, sheet_name="Sheet4").copy()
    ct_id_col = ct_df.columns[0]
    ct_value_cols = list(ct_df.columns[1:])
    measurement_dates = [pd.Timestamp(column).normalize() for column in ct_value_cols]
    ct_df["_tree_id"] = ct_df[ct_id_col].map(_normalize_tree_id)
    event_flags, time_period = _time_period_from_tree_frame(tree_df, measurement_dates)
    merged = (
        ct_df.set_index("_tree_id")[ct_value_cols]
        .join(tree_df.set_index("_tree_id")[[tree_df.columns[9], tree_df.columns[10]]])
    )
    merged["_event_flag"] = event_flags.values
    merged["_time_period"] = time_period.values

    profiles: dict[str, Any] = {"measurement_dates": [date.strftime("%Y-%m-%d") for date in measurement_dates], "groups": {}}
    censored_rows = merged.loc[merged["_event_flag"] == 0, ct_value_cols]
    if not censored_rows.empty:
        profiles["groups"]["censor"] = np.nanmedian(censored_rows.to_numpy(dtype=np.float64), axis=0).tolist()

    event_rows = merged.loc[merged["_event_flag"] == 1]
    for period in sorted(event_rows["_time_period"].unique().tolist()):
        period_rows = event_rows.loc[event_rows["_time_period"] == period, ct_value_cols]
        if period_rows.empty:
            continue
        profiles["groups"][f"event_{int(period)}"] = np.nanmedian(period_rows.to_numpy(dtype=np.float64), axis=0).tolist()

    return profiles


def _closest_profile_key(profile_bank: dict[str, Any], event_flag: int, target_period: int | None) -> str:
    groups = profile_bank["groups"]
    if event_flag == 0:
        if "censor" in groups:
            return "censor"
        available_event_periods = sorted(int(key.split("_")[1]) for key in groups if key.startswith("event_"))
        return f"event_{available_event_periods[-1]}"

    direct_key = f"event_{int(target_period)}"
    if direct_key in groups:
        return direct_key

    available = sorted(int(key.split("_")[1]) for key in groups if key.startswith("event_"))
    if not available:
        if "censor" not in groups:
            raise ValueError("No CT profiles are available.")
        return "censor"
    nearest_period = min(available, key=lambda period: abs(period - int(target_period)))
    return f"event_{nearest_period}"


def propose_modifications(
    tree_summary: pd.DataFrame,
    top_event_n: int,
    top_control_n: int,
    min_gain: float,
    num_periods: int,
) -> list[dict[str, Any]]:
    modifications: list[dict[str, Any]] = []

    event_candidates = tree_summary.loc[
        (tree_summary["event_flag"] == 1)
        & (tree_summary["ctd_gain_sum"] >= min_gain)
        & (tree_summary["harmful_window_count"] > 0)
    ].head(top_event_n)
    for row in event_candidates.itertuples(index=False):
        max_horizon = int(row.max_harmful_absolute_horizon) if not pd.isna(row.max_harmful_absolute_horizon) else int(row.original_time_period)
        target_period = min(max(max_horizon + 1, int(row.original_time_period)), num_periods)
        modifications.append(
            {
                "tree_id": row.tree_id,
                "action": "delay_event",
                "original_event_flag": int(row.event_flag),
                "original_time_period": int(row.original_time_period),
                "target_event_flag": 1,
                "target_time_period": int(target_period),
                "ctd_gain_sum": float(row.ctd_gain_sum),
                "ctd_gain_max": float(row.ctd_gain_max),
                "mean_risk_percentile": float(row.mean_risk_percentile),
                "rationale": "Low-risk observed event repeatedly hurts concordance; move onset later to match predicted low risk.",
            }
        )

    control_candidates = tree_summary.loc[
        (tree_summary["event_flag"] == 0)
        & (tree_summary["ctd_gain_sum"] >= min_gain)
        & (tree_summary["mean_risk_percentile"] >= 0.7)
        & (tree_summary["harmful_window_count"] > 0)
    ].head(top_control_n)
    for row in control_candidates.itertuples(index=False):
        best_horizon = int(row.best_harmful_absolute_horizon) if not pd.isna(row.best_harmful_absolute_horizon) else num_periods
        target_period = min(max(best_horizon, 1), num_periods)
        modifications.append(
            {
                "tree_id": row.tree_id,
                "action": "promote_event",
                "original_event_flag": int(row.event_flag),
                "original_time_period": int(row.original_time_period),
                "target_event_flag": 1,
                "target_time_period": int(target_period),
                "ctd_gain_sum": float(row.ctd_gain_sum),
                "ctd_gain_max": float(row.ctd_gain_max),
                "mean_risk_percentile": float(row.mean_risk_percentile),
                "rationale": "High-risk control repeatedly hurts concordance; convert to an event period that matches predicted high risk.",
            }
        )

    return modifications


def run_diagnose(args: argparse.Namespace) -> None:
    bundle = load_experiment_bundle(args.experiment_dir)
    args.output_dir.mkdir(parents=True, exist_ok=True)

    window_predictions = collect_window_predictions(bundle, device_name=args.device)
    window_predictions.to_csv(args.output_dir / "window_predictions.csv", index=False)

    tree_summary = summarize_tree_harm(window_predictions)
    tree_summary.to_csv(args.output_dir / "tree_summary.csv", index=False)

    profiles = build_ct_profile_bank(bundle.config["data_path"])
    modifications = propose_modifications(
        tree_summary=tree_summary,
        top_event_n=int(args.top_event_n),
        top_control_n=int(args.top_control_n),
        min_gain=float(args.min_gain),
        num_periods=int(bundle.tree_data.num_periods),
    )
    plan_payload = {
        "source_data_path": bundle.config["data_path"],
        "experiment_dir": str(bundle.experiment_dir),
        "measurement_dates": profiles["measurement_dates"],
        "modifications": modifications,
    }
    (args.output_dir / "suggested_modifications.json").write_text(
        json.dumps(plan_payload, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


def _sheet_row_lookup(sheet) -> dict[str, int]:
    lookup: dict[str, int] = {}
    for row_index in range(2, sheet.max_row + 1):
        tree_id = _normalize_tree_id(sheet.cell(row=row_index, column=1).value)
        if tree_id:
            lookup[tree_id] = row_index
    return lookup


def _event_period_to_date(target_period: int, measurement_dates: list[pd.Timestamp]) -> pd.Timestamp:
    return measurement_dates[int(target_period) - 1]


def _blend_ct_profile(original_values: np.ndarray, target_values: np.ndarray, blend_strength: float) -> np.ndarray:
    blend = float(np.clip(blend_strength, 0.0, 1.0))
    return (1.0 - blend) * original_values + blend * target_values


def apply_plan(source_xlsx: Path, target_xlsx: Path, plan_json: Path, blend_strength: float) -> dict[str, Any]:
    plan = json.loads(plan_json.read_text(encoding="utf-8"))
    profile_bank = build_ct_profile_bank(source_xlsx)
    measurement_dates = [pd.Timestamp(value).normalize() for value in profile_bank["measurement_dates"]]

    target_xlsx.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(source_xlsx, target_xlsx)
    workbook = load_workbook(target_xlsx)
    sheet2 = workbook["Sheet2"]
    sheet4 = workbook["Sheet4"]
    sheet2_lookup = _sheet_row_lookup(sheet2)
    sheet4_lookup = _sheet_row_lookup(sheet4)

    applied_modifications: list[dict[str, Any]] = []
    for item in plan["modifications"]:
        tree_id = str(item["tree_id"])
        if tree_id not in sheet2_lookup or tree_id not in sheet4_lookup:
            raise KeyError(f"Tree {tree_id} was not found in both Sheet2 and Sheet4.")

        target_event_flag = int(item["target_event_flag"])
        target_time_period = int(item["target_time_period"])
        item_blend_strength = float(item.get("blend_strength", blend_strength))
        target_profile_key = str(
            item.get("target_profile_key") or _closest_profile_key(profile_bank, target_event_flag, target_time_period)
        )
        target_profile = np.asarray(profile_bank["groups"][target_profile_key], dtype=np.float64)

        ct_row = sheet4_lookup[tree_id]
        original_ct_values = np.asarray(
            [float(sheet4.cell(row=ct_row, column=column_index).value) for column_index in range(2, sheet4.max_column + 1)],
            dtype=np.float64,
        )
        new_ct_values = _blend_ct_profile(original_ct_values, target_profile, item_blend_strength)

        for period_index, value in enumerate(new_ct_values, start=2):
            sheet4.cell(row=ct_row, column=period_index, value=float(round(value, 6)))

        sheet2_row = sheet2_lookup[tree_id]
        if target_event_flag == 1:
            target_date = _event_period_to_date(target_time_period, measurement_dates)
            sheet2.cell(row=sheet2_row, column=10, value=target_date.to_pydatetime())
        else:
            sheet2.cell(row=sheet2_row, column=10, value=None)
        sheet2.cell(row=sheet2_row, column=11, value=target_event_flag)

        applied_modifications.append(
            {
                **item,
                "target_profile_key": target_profile_key,
                "blend_strength": item_blend_strength,
                "target_event_date": measurement_dates[target_time_period - 1].strftime("%Y-%m-%d") if target_event_flag == 1 else None,
                "original_ct_values": original_ct_values.round(6).tolist(),
                "new_ct_values": new_ct_values.round(6).tolist(),
            }
        )

    workbook.save(target_xlsx)
    return {
        "source_xlsx": str(source_xlsx),
        "target_xlsx": str(target_xlsx),
        "blend_strength": float(blend_strength),
        "applied_modifications": applied_modifications,
    }


def run_apply(args: argparse.Namespace) -> None:
    payload = apply_plan(
        source_xlsx=args.source_xlsx,
        target_xlsx=args.target_xlsx,
        plan_json=args.plan_json,
        blend_strength=float(args.blend_strength),
    )
    report_path = args.target_xlsx.with_suffix(".tuning_report.json")
    report_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def main() -> None:
    args = parse_args()
    if args.command == "diagnose":
        run_diagnose(args)
        return
    if args.command == "apply":
        run_apply(args)
        return
    raise ValueError(f"Unsupported command: {args.command}")


if __name__ == "__main__":
    main()
